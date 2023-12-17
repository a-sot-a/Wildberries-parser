from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as xlImage
from io import BytesIO
from PIL import Image
import os
import requests

path = os.getcwd()
product_req = input('Какой товар искать будем?: ')

URL = 'https://search.wb.ru/exactmatch/ru/common/v4/search?'
HEADERS = {'Accept': "*/*", 'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
PARAMS = {
    'query': product_req,
    'resultset': 'catalog',
    'limit': 100,
    'sort': 'popular',
    'page': 1,
    'appType': 128,
    'curr': 'kzt',
    'lang': 'ru',
    'dest': -1257786,
    'spp': 27
}

response = requests.get(URL, headers=HEADERS, params=PARAMS)
data = response.json()

def create_workbook():
    # Создание книги и листа
    global wb, ws
    wb = Workbook()
    ws = wb.active

def set_column_widths():
    # Установка ширины столбцов
    font_size = 11
    cols_dict = {}

    for row in ws.rows:
        for cell in row:
            letter = cell.column_letter
            if cell.value:
                cell.font = Font(size=font_size)
                len_cell = len(str(cell.value))
                len_cell_dict = 0
                if letter in cols_dict:
                    len_cell_dict = cols_dict[letter]
                if len_cell > len_cell_dict:
                    cols_dict[letter] = len_cell
                    new_width_col = len_cell * font_size ** (font_size * 0.009)
                    ws.column_dimensions[cell.column_letter].width = new_width_col
    ws.column_dimensions["A"].width = int(200 * .2)

def add_headers():
    # Добавление заголовков
    cell_names = ['Картинка', 'Модель', 'Скидка', 'Цена без скидки', 'Цена со скидкой', 'Артикул', 'Рейтинг', 'Бренд']
    cell_coordinates = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']

    cell_dict = dict(zip(cell_names, cell_coordinates))

    for names, coordinates in cell_dict.items():
        ws[coordinates] = names
        ws[coordinates].fill = PatternFill('solid', fgColor="406fe6")

def center_text_in_rows():
    # Выравнивание текста в строках по центру
    set_column_widths()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def get_img(url, size=(100, 100)):
    # Получение изображения по URL
    r = requests.get(url, stream=True)

    r.raw.decode_content = True
    img = Image.open(r.raw)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)

def insert_row(ws, img_url, model, name, sale, price, salePrice, id, supplierRating, size=(200, 200)):
    # Вставка строки данных в таблицу
    img = xlImage(get_img(img_url, size=size))
    row_num = ws.max_row + 1
    cell_addr = f"A{row_num}"
    img.anchor = cell_addr
    ws.add_image(img)
    ws[f"B{row_num}"] = name
    ws[f"C{row_num}"] = str(sale) + ' %'
    ws[f"D{row_num}"] = str(price) + ' тг'
    ws[f"E{row_num}"] = str(salePrice) + ' тг'
    ws[f"F{row_num}"] = id
    ws[f"G{row_num}"] = str(supplierRating) + '★'
    ws[f"H{row_num}"] = model
    ws.row_dimensions[row_num].height = int(size[1] * .8)
    ws.column_dimensions["A"].width = int(size[0] * .2)

def insert_data(data):
    # Вставка данных из API в таблицу
    card_count = 1
    for i in data['data']['products']:
        try:
            for j in range(1, 99):
                if j < 10:
                    j = '0' + str(j)
                pic_url = f"https://basket-{j}.wbcontent.net/vol{i['id']//100000}/part{i['id']//1000}/{i['id']}/images/big/1.webp"

                response = requests.get(pic_url)

                if response.ok:
                    access_url = pic_url
                    size = (200, 200)
                    insert_row(ws, access_url, i['brand'], i['name'], i['sale'], i['priceU']//100, i['salePriceU']//100, i['id'], i['supplierRating'], size=size)
                    print(access_url)
                    print(card_count)
                    break
            card_count += 1
        except Exception as e:
            print(f"Что-то пошло не так. Ошибка: {e}")

def main():
    # Основная логика программы
    create_workbook()
    add_headers()
    insert_data(data)
    center_text_in_rows()
    wb.save(f'{product_req}.xlsx')

# Вызов основной функции
main()
